# Copyright (c) 2024-2026 Pikar AI. All rights reserved.
# Proprietary and confidential. See LICENSE file for details.

"""Shared MIME-aware document text extraction for vault ingestion.

This module centralises text extraction so both the user-facing Knowledge
Vault (``app/routers/vault.py``) and the admin knowledge service
(``app/services/knowledge_service.py``) can share the same truthful parsing
behaviour instead of maintaining separate, weaker ad-hoc decoders.

Exports:
    extract_text_from_bytes — Extract text from file bytes by MIME type.
                              Returns ``None`` for storage-only (non-searchable)
                              formats; raises ``ExtractionError`` on parse failure.
    is_searchable_format    — Return True when a MIME type maps to a format that
                              can be embedded as searchable text.
    ExtractionError         — Raised when extraction is attempted but fails.

Searchable formats (at minimum for v7):
    - text/plain
    - text/markdown, text/x-markdown, text/md
    - application/pdf                           (via pypdf)
    - application/vnd.openxmlformats-officedocument.wordprocessingml.document
                                                (via python-docx)
    - application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
                                                (via openpyxl)

All other MIME types (image/*, video/*, audio/*, and unknown binary formats
that do not resolve to text/PDF/DOCX/XLSX by MIME, filename, or OOXML sniffing)
are treated as storage-only and return ``None``.
"""

from __future__ import annotations

import io
import logging
import os
import zipfile
from typing import Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Lazy imports — keep module importable even if pypdf / docx are absent in
# some environments.  The names are module-level so tests can patch them.
# ---------------------------------------------------------------------------

try:
    import pypdf  # type: ignore[import]
except ImportError:  # pragma: no cover
    pypdf = None  # type: ignore[assignment]

try:
    import docx  # type: ignore[import]
except ImportError:  # pragma: no cover
    docx = None  # type: ignore[assignment]

try:
    from openpyxl import load_workbook  # type: ignore[import]
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]

# ---------------------------------------------------------------------------
# Public contract
# ---------------------------------------------------------------------------

#: MIME types that can be embedded as searchable text
_SEARCHABLE_MIMES: frozenset[str] = frozenset(
    [
        "text/plain",
        "text/markdown",
        "text/x-markdown",
        "text/md",
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ]
)

_TEXT_EXTENSIONS: frozenset[str] = frozenset(
    [
        ".txt",
        ".md",
        ".markdown",
        ".csv",
        ".tsv",
        ".json",
        ".py",
        ".js",
        ".ts",
        ".html",
        ".css",
        ".sql",
        ".xml",
        ".yaml",
        ".yml",
    ]
)

_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_DOC_LEGACY_MIME = "application/msword"
_XLS_LEGACY_MIME = "application/vnd.ms-excel"


class ExtractionError(Exception):
    """Raised when a supported format cannot be parsed (e.g. corrupt file)."""


def _normalise_mime(mime_type: str | None) -> str:
    """Strip charset / boundary suffixes and lowercase the MIME string."""
    if not mime_type:
        return ""
    return mime_type.lower().split(";")[0].strip()


def _normalise_extension(filename: str | None) -> str:
    """Return the lowercase suffix for *filename* or ``""`` when absent."""
    if not filename:
        return ""
    return os.path.splitext(filename)[1].lower()


def _detect_ooxml_family(file_bytes: bytes) -> str | None:
    """Return ``docx`` / ``xlsx`` when bytes look like an OOXML archive."""
    if not file_bytes.startswith(b"PK"):
        return None

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            names = archive.namelist()
    except zipfile.BadZipFile:
        return None

    if any(name.startswith("word/") for name in names):
        return "docx"
    if any(name.startswith("xl/") for name in names):
        return "xlsx"
    return None


def _resolve_extraction_target(
    file_bytes: bytes,
    mime_type: str | None,
    filename: str | None,
) -> str | None:
    """Resolve the best parser target for uploaded content."""
    normalised = _normalise_mime(mime_type)
    extension = _normalise_extension(filename)
    ooxml_family = _detect_ooxml_family(file_bytes)

    if normalised.startswith("text/") or extension in _TEXT_EXTENSIONS:
        return "text"

    if normalised == "application/pdf" or extension == ".pdf":
        return "pdf"

    if normalised == _DOCX_MIME or extension == ".docx":
        return "docx"

    if normalised == _XLSX_MIME or extension == ".xlsx":
        return "xlsx"

    if ooxml_family == "docx":
        return "docx"

    if ooxml_family == "xlsx":
        return "xlsx"

    if normalised == _DOC_LEGACY_MIME or extension == ".doc":
        return "legacy-doc"

    if normalised == _XLS_LEGACY_MIME or extension == ".xls":
        return "legacy-xls"

    return None


def is_searchable_format(
    mime_type: str | None,
    filename: str | None = None,
) -> bool:
    """Return ``True`` when *mime_type* maps to a searchable text format.

    Args:
        mime_type: Raw MIME type string (may include charset suffix).
        filename: Optional filename for extension-based fallback.

    Returns:
        ``True`` if the format can be embedded, ``False`` otherwise.
    """
    normalised = _normalise_mime(mime_type)
    extension = _normalise_extension(filename)

    if not normalised:
        return extension in _TEXT_EXTENSIONS or extension in {".pdf", ".docx", ".xlsx"}
    # Also accept any text/* sub-type
    if normalised.startswith("text/"):
        return True
    if normalised in _SEARCHABLE_MIMES:
        return True
    return extension in _TEXT_EXTENSIONS or extension in {".pdf", ".docx", ".xlsx"}


def extract_text_from_bytes(
    file_bytes: bytes,
    mime_type: str | None,
    *,
    filename: str | None = None,
) -> Optional[str]:
    """Extract searchable text from raw file bytes.

    Args:
        file_bytes: Raw binary content downloaded from storage.
        mime_type: MIME type string for the file (may include charset suffix).
        filename: Optional filename for extension-based fallback and clearer
            parser selection when browsers upload files as generic binary blobs.

    Returns:
        Extracted text string (may be empty ``""`` if the file has no text),
        or ``None`` if *mime_type* is not a searchable format (storage-only).

    Raises:
        ExtractionError: When the file is a supported searchable format but
            parsing fails (e.g. corrupt PDF, truncated DOCX).
    """
    target = _resolve_extraction_target(file_bytes, mime_type, filename)

    if target is None:
        return None

    if target == "legacy-doc":
        raise ExtractionError(
            "Legacy DOC extraction is not supported yet. "
            "Please upload a DOCX, PDF, or text export."
        )

    if target == "legacy-xls":
        raise ExtractionError(
            "Legacy XLS extraction is not supported yet. "
            "Please upload an XLSX or CSV export."
        )

    if target == "pdf":
        return _extract_pdf(file_bytes)

    if target == "docx":
        return _extract_docx(file_bytes)

    if target == "xlsx":
        return _extract_xlsx(file_bytes)

    return file_bytes.decode("utf-8", errors="replace")


# ---------------------------------------------------------------------------
# Internal extraction helpers
# ---------------------------------------------------------------------------


def _extract_pdf(file_bytes: bytes) -> str:
    """Extract text from PDF bytes via pypdf.

    Raises:
        ExtractionError: On any pypdf parse failure.
    """
    if pypdf is None:  # pragma: no cover
        raise ExtractionError("PDF extraction unavailable: pypdf not installed")
    try:
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        parts: list[str] = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                parts.append(text)
        return "\n".join(parts)
    except Exception as exc:
        raise ExtractionError(f"PDF extraction failed: {exc}") from exc


def _extract_docx(file_bytes: bytes) -> str:
    """Extract text from DOCX bytes via python-docx.

    Raises:
        ExtractionError: On any python-docx parse failure.
    """
    if docx is None:  # pragma: no cover
        raise ExtractionError("DOCX extraction unavailable: python-docx not installed")
    try:
        doc = docx.Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text)
    except Exception as exc:
        raise ExtractionError(f"DOCX extraction failed: {exc}") from exc


def _extract_xlsx(file_bytes: bytes) -> str:
    """Extract text-like tabular content from XLSX bytes via openpyxl."""
    if load_workbook is None:  # pragma: no cover
        raise ExtractionError("XLSX extraction unavailable: openpyxl not installed")

    workbook = None
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=True)
        sheets: list[str] = []

        for worksheet in workbook.worksheets:
            rows: list[str] = []
            for row in worksheet.iter_rows(values_only=True):
                cells = [
                    str(cell).strip()
                    for cell in row
                    if cell is not None and str(cell).strip()
                ]
                if cells:
                    rows.append("\t".join(cells))

            if rows:
                sheets.append(f"[Sheet: {worksheet.title}]\n" + "\n".join(rows))

        return "\n\n".join(sheets)
    except Exception as exc:
        raise ExtractionError(f"XLSX extraction failed: {exc}") from exc
    finally:
        if workbook is not None and hasattr(workbook, "close"):
            workbook.close()
